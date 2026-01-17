import pandas as pd
import numpy as np
from datetime import datetime, timedelta, timezone
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from flask import Flask, request, send_file, jsonify
import io

app = Flask(__name__)

# ---------------------------------------------------------
# SHARED UTILITIES
# ---------------------------------------------------------

def get_malaysia_timestamp():
    """Returns formatted Malaysia timestamp."""
    malaysia_time = datetime.now(timezone.utc) + timedelta(hours=8)
    return malaysia_time.strftime("%d-%b-%Y %I:%M %p")

def generate_time_slots():
    """Generates 30-minute time slots from 08:05 to 17:35."""
    start_time = datetime.strptime("08:05:00", "%H:%M:%S")
    end_time = datetime.strptime("17:35:00", "%H:%M:%S")
    time_slots = []
    current = start_time
    while current <= end_time:
        time_slots.append(current.strftime("%H:%M:%S"))
        current += timedelta(minutes=30)
    return time_slots

def load_and_clean_data(file_stream):
    """Reads Excel stream, merges sheets, and cleans columns."""
    try:
        all_sheets = pd.read_excel(file_stream, header=3, sheet_name=None)
    except Exception as e:
        raise ValueError(f"Invalid Excel file: {str(e)}")

    df = pd.concat(all_sheets.values(), ignore_index=True)

    # Column Mapping
    if 'DAY' in df.columns and 'DAY.1' in df.columns:
        df = df.drop(columns=['DAY'])
        df = df.rename(columns={'DAY.1': 'DAY'})

    df.columns = df.columns.str.strip().str.upper()
    
    # Required Columns
    required_cols = ['DAY', 'START TIME', 'SUBJECT', 'TEACHER', 'ROOM', 'GROUP', 'INTAKE']
    missing_cols = [col for col in required_cols if col not in df.columns]
    
    if 'START TIME' not in df.columns:
        raise ValueError(f"Missing critical column: START TIME. Found: {list(df.columns)}")

    # Filter & Clean
    existing_cols = [c for c in required_cols if c in df.columns]
    df = df[existing_cols].copy()
    
    # Standardize Time
    try:
        df['START TIME'] = pd.to_datetime(df['START TIME'].astype(str).str.strip(), format='%H:%M:%S').dt.strftime('%H:%M:%S')
    except ValueError:
        df['START TIME'] = pd.to_datetime(df['START TIME'], errors='coerce').dt.strftime('%H:%M:%S')
    
    df = df.dropna(subset=['START TIME'])

    # Clean Text Columns
    text_cols = ['ROOM', 'GROUP', 'TEACHER', 'SUBJECT', 'INTAKE']
    for col in text_cols:
        df[col] = df.get(col, pd.Series([""] * len(df))).astype(str).str.strip()

    # Create Group Column
    df['INTAKE_GROUP'] = df['INTAKE'] + " " + df['GROUP']
    
    return df

def format_worksheet(worksheet, timestamp_text, first_col_width=25):
    """Applies standard formatting to the output worksheet."""
    # 1. Header Timestamp
    worksheet['A1'] = timestamp_text
    worksheet.merge_cells('A1:D1')
    worksheet['A1'].font = Font(size=11, bold=True, italic=True, color="555555")

    # 2. Styles
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    header_col_font = Font(size=14, bold=True, name='Calibri')
    regular_font = Font(size=11, name='Calibri')

    # 3. Apply Styles
    for row in worksheet.iter_rows(min_row=2):
        for col_idx, cell in enumerate(row):
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
            cell.border = thin_border
            
            if col_idx == 0: # First Column (Teacher or Room Name)
                cell.font = header_col_font
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            else:
                cell.font = regular_font

    # 4. Column Widths
    worksheet.column_dimensions['A'].width = first_col_width
    for i in range(2, worksheet.max_column + 1):
        col_letter = get_column_letter(i)
        worksheet.column_dimensions[col_letter].width = 22

# ---------------------------------------------------------
# GENERATION LOGIC
# ---------------------------------------------------------

def generate_schedule_excel(df, view_type='room'):
    """
    Generates the Excel binary based on view_type ('room' or 'teacher').
    """
    output_buffer = io.BytesIO()
    time_slots = generate_time_slots()
    header_text = f"File Generated on: {get_malaysia_timestamp()}"

    # Grouping Logic
    df_grouped = df.groupby(['DAY', 'START TIME', 'SUBJECT', 'TEACHER', 'ROOM'])['INTAKE_GROUP'].apply(lambda x: '\n'.join(x.unique())).reset_index()

    day_rank = {
        'MON': 0, 'MONDAY': 0, 'TUE': 1, 'TUESDAY': 1,
        'WED': 2, 'WEDNESDAY': 2, 'THU': 3, 'THURSDAY': 3,
        'FRI': 4, 'FRIDAY': 4, 'SAT': 5, 'SATURDAY': 5,
        'SUN': 6, 'SUNDAY': 6
    }
    unique_days = sorted(df_grouped['DAY'].unique(), key=lambda d: day_rank.get(str(d).strip().upper(), 99))

    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        for day in unique_days:
            day_data = df_grouped[df_grouped['DAY'] == day]
            if day_data.empty: continue

            # --- CONFIGURATION BASED ON VIEW TYPE ---
            if view_type == 'room':
                primary_col = 'ROOM'
                display_cols = ['TEACHER', 'SUBJECT', 'INTAKE_GROUP']
            else:  # teacher view
                primary_col = 'TEACHER'
                display_cols = ['ROOM', 'SUBJECT', 'INTAKE_GROUP']

            # Filter valid rows
            unique_primary = sorted([x for x in day_data[primary_col].unique() if x != "nan" and x != ""])
            if not unique_primary: continue

            # Initialize Matrix
            matrix_cols = [primary_col] + time_slots
            matrix_df = pd.DataFrame(index=unique_primary, columns=matrix_cols)
            matrix_df[primary_col] = matrix_df.index

            # Fill Matrix
            for _, row in day_data.iterrows():
                primary_val = row[primary_col]
                if primary_val not in matrix_df.index: continue

                start_str = row['START TIME']
                
                # Construct Cell Text
                cell_text_parts = [row[col] for col in display_cols]
                cell_text = "\n".join(cell_text_parts)

                if start_str in matrix_cols:
                    # Update function
                    def update_cell(r, c, text):
                        if pd.isna(matrix_df.at[r, c]):
                            matrix_df.at[r, c] = text
                        else:
                            matrix_df.at[r, c] += f"\n---\n{text}"

                    update_cell(primary_val, start_str, cell_text)
                    
                    # Fill next 30 min slot (simple 1-hour assumption)
                    start_idx = matrix_cols.index(start_str)
                    if start_idx + 1 < len(matrix_cols):
                        next_slot = matrix_cols[start_idx + 1]
                        update_cell(primary_val, next_slot, cell_text)

            # Write Sheet
            final_df = matrix_df.reset_index(drop=True)
            sheet_name_clean = str(day).split(' ')[0].upper()[:30]
            final_df.to_excel(writer, sheet_name=sheet_name_clean, index=False, startrow=1)

            # Format Sheet
            if sheet_name_clean in writer.book.sheetnames:
                format_worksheet(writer.book[sheet_name_clean], header_text)

    output_buffer.seek(0)
    return output_buffer

# ---------------------------------------------------------
# API ROUTES
# ---------------------------------------------------------

@app.route('/api/schedule/room', methods=['POST'])
def schedule_room():
    return handle_request(view_type='room', default_name='Room_Schedule.xlsx')

@app.route('/api/schedule/teacher', methods=['POST'])
def schedule_teacher():
    return handle_request(view_type='teacher', default_name='Teacher_Schedule.xlsx')

def handle_request(view_type, default_name):
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400

    output_filename = request.form.get('filename', default_name)
    if not output_filename.endswith('.xlsx'):
        output_filename += '.xlsx'

    try:
        # Process
        df = load_and_clean_data(file)
        output_buffer = generate_schedule_excel(df, view_type=view_type)

        return send_file(
            output_buffer,
            as_attachment=True,
            download_name=output_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except ValueError as ve:
        return jsonify({"error": str(ve)}), 400
    except Exception as e:
        return jsonify({"error": "Internal Server Error", "details": str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5001, host='0.0.0.0')